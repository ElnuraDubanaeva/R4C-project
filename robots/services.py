import os
from string import ascii_uppercase
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl import Workbook
from django.db.models import Count
from django.conf import settings
from common.constants import EXCEL_ROBOT_HEADERS
from .models import Robot


class ExcelServices:
    headers = EXCEL_ROBOT_HEADERS
    __model = Robot

    @staticmethod
    def create_folder(folder_name: str) -> None:
        """function for create folder if not exists"""
        try:
            os.mkdir(f"{settings.MEDIA_ROOT}")
            os.mkdir(f"{settings.MEDIA_ROOT}/{folder_name}")
        except FileExistsError:
            pass

    @classmethod
    def style_excel(cls, worksheet):
        letters = ascii_uppercase[: len(cls.headers)]
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            for column_index, cell_value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.alignment = Alignment(horizontal="justify")

        worksheet.column_dimensions["A"].width = 10
        for i in letters[1:]:
            worksheet.column_dimensions[i].width = 20
            cell = worksheet[f"{i}1"]
            cell.font = Font(bold=True)
            side = Side(border_style="thick")
            cell.border = Border(top=side, bottom=side, left=side, right=side)
        worksheet.column_dimensions["D"].width = 25

    @classmethod
    def get_robots_data(cls, robot_model):
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        robots = cls.__model.objects.filter(
            created_at__range=(start_date, end_date),
            is_deleted=False,
            model=robot_model,
        )
        robot_serials = list(robots.values_list("serial", flat=True))

        robots_query = {}
        data = [cls.headers]
        distinct_robots = robots.order_by("serial").distinct("serial")
        for index, robot in enumerate(distinct_robots):
            count = robot_serials.count(robot.serial)
            data.append([index + 1, robot.model, robot.version, count])
            robots_query[robot.model] = data
        return data

    @classmethod
    def get_excel_robots(cls):
        cls.create_folder("excel")
        output_file_path = os.path.join(settings.MEDIA_ROOT, "excel", "robots.xlsx")
        workbook = Workbook()

        default_sheet = workbook.active
        workbook.remove(default_sheet)

        unique_models = list(
            cls.__model.objects.order_by("model")
            .distinct("model")
            .values_list("model", flat=True)
        )

        for model in unique_models:
            worksheet = workbook.create_sheet(title=model)
            data = cls.get_robots_data(model)

            for row_num, row_data in enumerate(data, start=1):
                for col_num, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = value

            cls.style_excel(worksheet=worksheet)

        workbook.save(output_file_path)
        return workbook
